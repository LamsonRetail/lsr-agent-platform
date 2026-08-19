"""Giám sát chỉ tiêu ngày tồn kho (DOI) & cảnh báo an toàn cho cuộc họp S&OP.

MVP: đọc dữ liệu từ file Excel (theo config/thresholds.yaml). Sau này thay
nguồn "excel" bằng "bigquery" mà không đổi phần tính toán/cảnh báo phía dưới —
xem `load_source_aggregate`.

Cách tính ngày tồn kho công ty (aggregate) từ dữ liệu SKU:
    ngày_tồn_kho = SUM(số lượng các SKU) / SUM(tốc độ bán các SKU)
(không lấy trung bình cộng NTK từng SKU — sẽ bị lệch bởi SKU nhỏ/TĐB=0)

Chạy:
    python inventory_days.py --config ../config/thresholds.yaml \
        --excel "C:/.../QL KE HOACH HANG HOA HAPAS_THEO DOI TON KHO_DATA GOC.xlsx"
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass

import openpyxl
import requests
import yaml
from dotenv import find_dotenv, load_dotenv


@dataclass
class SourceResult:
    key: str
    label: str
    current_qty: float | None
    total_qty: float | None
    velocity_per_day: float | None
    current_days: float | None
    total_days: float | None
    target_current: float
    target_total: float
    safety_current: float
    safety_total: float
    overstock_multiplier: float
    missing_data: bool = False

    def status(self, tier: str) -> str:
        """Trả về mức cảnh báo cho 1 tier ('current' hoặc 'total')."""
        if self.missing_data:
            return "missing"
        days = self.current_days if tier == "current" else self.total_days
        safety = self.safety_current if tier == "current" else self.safety_total
        target = self.target_current if tier == "current" else self.target_total
        if days is None:
            return "missing"
        if days <= safety:
            return "red"        # dưới ngưỡng an toàn -> đặt ngay, ngoài kế hoạch
        if days < target:
            return "yellow"     # dưới target nhưng còn an toàn
        if days >= target * self.overstock_multiplier:
            return "overstock"  # vượt target quá xa -> tồn dư
        return "green"

    def suggested_order_qty(self, tier: str) -> float | None:
        """Số lượng cần đặt thêm để đưa ngày tồn kho (tier) về đúng target."""
        if self.missing_data or self.velocity_per_day in (None, 0):
            return None
        qty = self.current_qty if tier == "current" else self.total_qty
        target = self.target_current if tier == "current" else self.target_total
        if qty is None:
            return None
        gap_days = target - (qty / self.velocity_per_day)
        if gap_days <= 0:
            return None
        return round(gap_days * self.velocity_per_day)


def _resolve_column_index(header_row: list, name: str, *, use_last: bool = True) -> int:
    """Trả về index (1-based) của cột theo tên header.

    File nguồn có vài cột trùng tên (bản "tên cũ"/"tên mới" trước cột chính
    thức) — mặc định lấy cột **cuối cùng** khớp tên vì đó là cột đã hợp nhất.
    """
    matches = [i + 1 for i, v in enumerate(header_row) if v == name]
    if not matches:
        raise KeyError(f"Không tìm thấy cột '{name}' trong file excel")
    return matches[-1] if use_last else matches[0]


def load_excel_aggregate(source_key: str, cfg: dict) -> dict:
    """Đọc + cộng dồn số liệu SKU từ excel cho 1 nguồn theo config['excel']."""
    excel_cfg = cfg["excel"]
    path = cfg["_excel_path"]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[excel_cfg["sheet"]]

    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    velocity_idx = _resolve_column_index(header_row, excel_cfg["velocity_column"])
    current_idxs = [_resolve_column_index(header_row, n) for n in excel_cfg["current_qty_columns"]]
    total_idx = _resolve_column_index(header_row, excel_cfg["total_qty_column"])

    sum_current = 0.0
    sum_total = 0.0
    sum_velocity = 0.0
    rows_used = 0

    for r in range(2, ws.max_row + 1):
        sku = ws.cell(row=r, column=2).value
        if sku is None:
            continue  # dòng rỗng/dòng ghi chú cuối file
        velocity = ws.cell(row=r, column=velocity_idx).value
        velocity = _to_number(velocity)
        if velocity is None:
            continue  # không có tốc độ bán -> loại khỏi mẫu số (tránh chia lệch)
        sum_velocity += velocity
        for idx in current_idxs:
            sum_current += _to_number(ws.cell(row=r, column=idx).value) or 0.0
        sum_total += _to_number(ws.cell(row=r, column=total_idx).value) or 0.0
        rows_used += 1

    return {
        "current_qty": sum_current,
        "total_qty": sum_total,
        "velocity_per_day": sum_velocity,
        "rows_used": rows_used,
    }


@dataclass
class SkuResult:
    """Ngày tồn kho của 1 mã SKU."""

    sku: str
    name: str
    collection: str | None      # BST
    current_qty: float
    total_qty: float
    velocity_per_day: float
    current_days: float | None  # None = không bán (TĐB=0) -> không quy đổi được ra ngày
    total_days: float | None


def load_excel_skus(cfg: dict) -> list[SkuResult]:
    """Đọc chi tiết từng SKU (không cộng dồn) để xếp hạng tồn cao/tồn thấp."""
    excel_cfg = cfg["excel"]
    wb = openpyxl.load_workbook(cfg["_excel_path"], data_only=True)
    ws = wb[excel_cfg["sheet"]]

    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    velocity_idx = _resolve_column_index(header_row, excel_cfg["velocity_column"])
    current_idxs = [_resolve_column_index(header_row, n) for n in excel_cfg["current_qty_columns"]]
    total_idx = _resolve_column_index(header_row, excel_cfg["total_qty_column"])

    skus: list[SkuResult] = []
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(row=r, column=2).value
        if sku is None:
            continue
        velocity = _to_number(ws.cell(row=r, column=velocity_idx).value)
        if velocity is None:
            continue
        current_qty = sum((_to_number(ws.cell(row=r, column=i).value) or 0.0) for i in current_idxs)
        total_qty = _to_number(ws.cell(row=r, column=total_idx).value) or 0.0
        skus.append(SkuResult(
            sku=str(sku),
            name=str(ws.cell(row=r, column=1).value or ""),
            collection=ws.cell(row=r, column=4).value,
            current_qty=current_qty,
            total_qty=total_qty,
            velocity_per_day=velocity,
            current_days=(current_qty / velocity) if velocity else None,
            total_days=(total_qty / velocity) if velocity else None,
        ))
    return skus


def render_sku_ranking(skus: list[SkuResult], cfg_source: SourceResult, *, top_n: int = 10) -> str:
    """Bảng SKU tồn cao nhất / tồn thấp nhất theo ngày tồn kho hiện tại."""
    rated = [s for s in skus if s.current_days is not None]
    # SKU còn hàng nhưng không bán được (TĐB=0) — không có ngày tồn kho, tách riêng.
    dead = [s for s in skus if s.current_days is None and s.current_qty > 0]

    high = sorted(rated, key=lambda s: s.current_days, reverse=True)[:top_n]
    # Tồn thấp: chỉ xét SKU đang thực sự bán (TĐB > 0) để loại nhiễu.
    low = sorted([s for s in rated if s.velocity_per_day > 0], key=lambda s: s.current_days)[:top_n]

    lines = ["", f"## Top {top_n} SKU **tồn cao** (ngày tồn hiện tại lớn nhất)", ""]
    lines.append("| SKU | Tên | BST | Tồn hiện tại | TĐB/ngày | Ngày tồn |")
    lines.append("|---|---|---|---|---|---|")
    for s in high:
        lines.append(f"| {s.sku} | {s.name} | {s.collection or '—'} | {s.current_qty:,.0f} | "
                      f"{s.velocity_per_day:,.1f} | **{s.current_days:,.0f}** |")

    lines += ["", f"## Top {top_n} SKU **tồn thấp** (rủi ro hết hàng)", ""]
    lines.append("| SKU | Tên | BST | Tồn hiện tại | TĐB/ngày | Ngày tồn |")
    lines.append("|---|---|---|---|---|---|")
    for s in low:
        lines.append(f"| {s.sku} | {s.name} | {s.collection or '—'} | {s.current_qty:,.0f} | "
                      f"{s.velocity_per_day:,.1f} | **{s.current_days:,.1f}** |")

    if dead:
        lines += ["", f"⚠️ **{len(dead)} SKU còn tồn nhưng TĐB = 0** (không bán được, không quy đổi ra ngày tồn): "
                      + ", ".join(s.sku for s in dead[:15])
                      + (f" … (+{len(dead) - 15} mã)" if len(dead) > 15 else "")]
    return "\n".join(lines)


def _to_number(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def load_source_aggregate(source_key: str, cfg: dict) -> dict | None:
    """Điểm mở rộng: thêm nguồn (vd 'bigquery') chỉ cần thêm nhánh ở đây."""
    data_source = cfg.get("data_source")
    if data_source == "excel":
        return load_excel_aggregate(source_key, cfg)
    return None  # nguồn chưa cấu hình (vd chưa có file MateMade VN/Hapas TL/NVL)


def evaluate(config_path: str, excel_path: str | None) -> list[SourceResult]:
    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    default_overstock_multiplier = config.get("overstock_multiplier", 1.5)

    results = []
    for key, cfg in config["sources"].items():
        cfg = dict(cfg)
        cfg["_excel_path"] = excel_path
        overstock_multiplier = cfg.get("overstock_multiplier", default_overstock_multiplier)
        agg = load_source_aggregate(key, cfg) if excel_path or cfg.get("data_source") != "excel" else None

        if agg is None:
            results.append(SourceResult(
                key=key, label=cfg["label"],
                current_qty=None, total_qty=None, velocity_per_day=None,
                current_days=None, total_days=None,
                target_current=cfg["target_days"]["current"],
                target_total=cfg["target_days"]["total"],
                safety_current=cfg["safety_threshold_days"]["current"],
                safety_total=cfg["safety_threshold_days"]["total"],
                overstock_multiplier=overstock_multiplier,
                missing_data=True,
            ))
            continue

        velocity = agg["velocity_per_day"]
        current_days = agg["current_qty"] / velocity if velocity else None
        total_days = agg["total_qty"] / velocity if velocity else None

        results.append(SourceResult(
            key=key, label=cfg["label"],
            current_qty=agg["current_qty"], total_qty=agg["total_qty"],
            velocity_per_day=velocity,
            current_days=current_days, total_days=total_days,
            target_current=cfg["target_days"]["current"],
            target_total=cfg["target_days"]["total"],
            safety_current=cfg["safety_threshold_days"]["current"],
            safety_total=cfg["safety_threshold_days"]["total"],
            overstock_multiplier=overstock_multiplier,
        ))
    return results


STATUS_ICON = {"green": "🟢", "yellow": "🟡", "red": "🔴", "overstock": "🟣", "missing": "⚪"}


def render_report(results: list[SourceResult]) -> str:
    lines = ["# Chỉ tiêu ngày tồn kho — báo cáo S&OP", ""]
    lines.append("| Nguồn | Hiện tại (ngày) | Target | An toàn | Tổng (ngày) | Target | An toàn | Đề xuất đặt thêm |")
    lines.append("|---|---|---|---|---|---|---|---|")
    for r in results:
        if r.missing_data:
            lines.append(f"| {r.label} | ⚪ chưa có dữ liệu | {r.target_current} | ≤{r.safety_current} | "
                          f"⚪ chưa có dữ liệu | {r.target_total} | ≤{r.safety_total} | — |")
            continue

        cur_icon = STATUS_ICON[r.status("current")]
        tot_icon = STATUS_ICON[r.status("total")]
        suggestion_parts = []
        for tier, icon in (("current", cur_icon), ("total", tot_icon)):
            if icon == "🔴":
                qty = r.suggested_order_qty(tier)
                if qty:
                    tier_label = "hiện tại" if tier == "current" else "tổng"
                    suggestion_parts.append(f"+{qty:,.0f} (để {tier_label} về target)")
        suggestion = "; ".join(suggestion_parts) if suggestion_parts else "—"

        lines.append(
            f"| {r.label} | {cur_icon} {r.current_days:.1f} | {r.target_current} | ≤{r.safety_current} | "
            f"{tot_icon} {r.total_days:.1f} | {r.target_total} | ≤{r.safety_total} | {suggestion} |"
        )

    reds = [r for r in results if not r.missing_data and ("red" in (r.status("current"), r.status("total")))]
    overstocked = [r for r in results if not r.missing_data and ("overstock" in (r.status("current"), r.status("total")))]
    lines.append("")
    if reds:
        lines.append(f"⚠️ **{len(reds)} nguồn dưới ngưỡng an toàn — cần đặt hàng ngay (ngoài kế hoạch):** "
                      + ", ".join(r.label for r in reds))
    else:
        lines.append("✅ Không có nguồn nào dưới ngưỡng an toàn.")
    if overstocked:
        lines.append(f"🟣 **{len(overstocked)} nguồn tồn dư (≥ target × {results[0].overstock_multiplier:g}):** "
                      + ", ".join(r.label for r in overstocked))
    return "\n".join(lines)


def render_lark_sku_block(skus: list[SkuResult], label: str, *, top_n: int = 5) -> str:
    """Khối SKU tồn cao/tồn thấp dạng text gọn cho tin Lark."""
    rated = [s for s in skus if s.current_days is not None]
    dead = [s for s in skus if s.current_days is None and s.current_qty > 0]
    high = sorted(rated, key=lambda s: s.current_days, reverse=True)[:top_n]
    low = sorted([s for s in rated if s.velocity_per_day > 0], key=lambda s: s.current_days)[:top_n]

    lines = ["", f"— Chi tiết SKU ({label}) —", "", f"🟣 Tồn cao nhất (top {top_n}):"]
    for s in high:
        lines.append(f"  • {s.sku} — {s.name}: {s.current_days:,.0f} ngày "
                      f"(tồn {s.current_qty:,.0f}, bán {s.velocity_per_day:,.1f}/ngày)")
    lines += ["", f"🔴 Tồn thấp nhất (top {top_n}):"]
    for s in low:
        lines.append(f"  • {s.sku} — {s.name}: {s.current_days:,.1f} ngày "
                      f"(tồn {s.current_qty:,.0f}, bán {s.velocity_per_day:,.1f}/ngày)")
    if dead:
        lines += ["", f"⚠️ {len(dead)} SKU còn tồn nhưng không bán được (TĐB=0) — vốn chết, "
                       f"không tính được ngày tồn."]
    return "\n".join(lines)


def render_lark_text(results: list[SourceResult], sku_blocks: list[str] | None = None) -> str:
    """Bản tóm tắt dạng text thuần (không markdown table) để gửi tin Lark."""
    lines = ["📦 Chỉ tiêu ngày tồn kho — báo cáo S&OP", ""]
    for r in results:
        if r.missing_data:
            lines.append(f"⚪ {r.label}: chưa có dữ liệu")
            continue
        cur_icon = STATUS_ICON[r.status("current")]
        tot_icon = STATUS_ICON[r.status("total")]
        line = (f"{r.label}: hiện tại {cur_icon} {r.current_days:.1f}j (target {r.target_current}, "
                f"an toàn ≤{r.safety_current}) · tổng {tot_icon} {r.total_days:.1f}j "
                f"(target {r.target_total}, an toàn ≤{r.safety_total})")
        for tier, icon in (("current", cur_icon), ("total", tot_icon)):
            if icon == "🔴":
                qty = r.suggested_order_qty(tier)
                if qty:
                    tier_label = "hiện tại" if tier == "current" else "tổng"
                    line += f" — đề xuất đặt thêm {qty:,.0f} ({tier_label})"
        lines.append(line)

    reds = [r for r in results if not r.missing_data and ("red" in (r.status("current"), r.status("total")))]
    overstocked = [r for r in results if not r.missing_data and ("overstock" in (r.status("current"), r.status("total")))]
    lines.append("")
    if reds:
        lines.append("⚠️ Cần đặt hàng ngay (ngoài kế hoạch): " + ", ".join(r.label for r in reds))
    if overstocked:
        lines.append("🟣 Tồn dư: " + ", ".join(r.label for r in overstocked))
    if not reds and not overstocked:
        lines.append("✅ Tất cả nguồn có dữ liệu đều trong ngưỡng an toàn/target.")
    if sku_blocks:
        lines.extend(sku_blocks)
    return "\n".join(lines)


def send_to_lark(text: str, chat_id: str, *, app_id: str, app_secret: str, domain: str) -> None:
    """Gửi tin nhắn text tới 1 nhóm Lark qua Custom App (tenant_access_token)."""
    resp = requests.post(
        f"{domain}/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": app_id, "app_secret": app_secret},
        timeout=30,
    )
    resp.raise_for_status()
    payload = resp.json()
    if payload.get("code") != 0:
        raise RuntimeError(f"Lấy tenant_access_token thất bại: {payload}")
    token = payload["tenant_access_token"]

    resp = requests.post(
        f"{domain}/open-apis/im/v1/messages",
        params={"receive_id_type": "chat_id"},
        headers={"Authorization": f"Bearer {token}"},
        json={"receive_id": chat_id, "msg_type": "text", "content": _text_content_json(text)},
        timeout=30,
    )
    resp.raise_for_status()
    payload = resp.json()
    if payload.get("code") != 0:
        raise RuntimeError(f"Gửi tin nhắn Lark thất bại: {payload}")


def _text_content_json(text: str) -> str:
    import json
    return json.dumps({"text": text}, ensure_ascii=False)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", default="../config/thresholds.yaml")
    parser.add_argument("--excel", required=True, help="Đường dẫn file excel data gốc (dùng cho nguồn có data_source: excel)")
    parser.add_argument("--lark-chat-id", help="Nếu truyền, gửi báo cáo vào nhóm Lark này (cần LARK_APP_ID/LARK_APP_SECRET trong .env)")
    parser.add_argument("--sku-detail", action="store_true", help="Kèm bảng SKU tồn cao / tồn thấp")
    parser.add_argument("--top-n", type=int, default=10, help="Số SKU hiển thị mỗi chiều (mặc định 10)")
    args = parser.parse_args()

    load_dotenv(find_dotenv(usecwd=True))

    results = evaluate(args.config, args.excel)
    print(render_report(results))

    sku_blocks: list[str] = []
    if args.sku_detail:
        with open(args.config, "r", encoding="utf-8") as f:
            config = yaml.safe_load(f)
        for key, cfg in config["sources"].items():
            if cfg.get("data_source") != "excel":
                continue
            cfg = dict(cfg)
            cfg["_excel_path"] = args.excel
            source = next(r for r in results if r.key == key)
            skus = load_excel_skus(cfg)
            print(f"\n---\n# Chi tiết SKU — {source.label}")
            print(render_sku_ranking(skus, source, top_n=args.top_n))
            # Tin Lark dùng bản gọn hơn (top 5) để không quá dài.
            sku_blocks.append(render_lark_sku_block(skus, source.label, top_n=min(5, args.top_n)))

    if args.lark_chat_id:
        # Ưu tiên bot riêng của agent này; fallback về app Lark mặc định của repo.
        app_id = os.environ.get("LARK_APP_ID_INVENTORY") or os.environ.get("LARK_APP_ID")
        app_secret = os.environ.get("LARK_APP_SECRET_INVENTORY") or os.environ.get("LARK_APP_SECRET")
        domain = os.environ.get("LARK_DOMAIN", "https://open.larksuite.com")
        if not app_id or not app_secret:
            print("\n[Lark] Thiếu LARK_APP_ID_INVENTORY/LARK_APP_SECRET_INVENTORY trong .env — bỏ qua gửi tin.", file=sys.stderr)
            return 1
        send_to_lark(render_lark_text(results, sku_blocks), args.lark_chat_id,
                     app_id=app_id, app_secret=app_secret, domain=domain)
        print(f"\n[Lark] Đã gửi báo cáo vào chat_id={args.lark_chat_id}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
